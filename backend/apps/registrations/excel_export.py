"""
Excel Export generator for JTC Registrations, Students, and Financial Data using openpyxl.
"""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from .models import Registration, RegistrationEvent, Participant
from apps.events.models import Event


def generate_registrations_workbook(qs=None) -> bytes:
    """
    Generates a professionally styled multi-tab Excel spreadsheet containing:
    1. Registrations & Payment Ledger
    2. Event Roster & Team Breakdown
    3. Category & Financial Overview
    """
    if qs is None:
        qs = Registration.objects.select_related('participant', 'participant__school', 'participant__group', 'payment_verified_by').prefetch_related('registration_events__event').all()

    wb = openpyxl.Workbook()

    # Define color palettes and typography
    header_fill = PatternFill(start_color='010B1E', end_color='010B1E', fill_type='solid')
    header_font = Font(name='Arial', size=11, bold=True, color='F0C040')  # Gold text on Dark Navy
    sub_header_fill = PatternFill(start_color='0F1F38', end_color='0F1F38', fill_type='solid')
    sub_header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')

    title_font = Font(name='Arial', size=14, bold=True, color='010B1E')
    meta_font = Font(name='Arial', size=9, italic=True, color='555555')
    bold_font = Font(name='Arial', size=10, bold=True, color='000000')
    regular_font = Font(name='Arial', size=10, color='111111')

    # Status Fills & Fonts
    status_styles = {
        'VERIFIED': {
            'fill': PatternFill(start_color='D1E7DD', end_color='D1E7DD', fill_type='solid'),
            'font': Font(name='Arial', size=10, bold=True, color='0F5132'),
        },
        'PENDING': {
            'fill': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
            'font': Font(name='Arial', size=10, bold=True, color='664D03'),
        },
        'REJECTED': {
            'fill': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
            'font': Font(name='Arial', size=10, bold=True, color='842029'),
        },
        'REFUNDED': {
            'fill': PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid'),
            'font': Font(name='Arial', size=10, bold=True, color='41464B'),
        },
    }

    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    # -------------------------------------------------------------
    # TAB 1: ALL REGISTRATIONS & PAYMENT LEDGER
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Student Registrations"
    ws1.views.sheetView[0].showGridLines = True

    # Title & Metadata
    ws1['A1'] = "JOSEPHITE TECH CLUB — SJIS TECH CARNIVAL 2026"
    ws1['A1'].font = title_font
    ws1['A2'] = f"Student Registrations & Payment Ledger • Generated: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')}"
    ws1['A2'].font = meta_font

    headers_1 = [
        "Reg ID",
        "Confirmation Code",
        "Registered At",
        "Student Name",
        "Email Address",
        "Phone / WhatsApp",
        "Institution / School",
        "Grade Level",
        "Academic Group",
        "Total Fee (BDT)",
        "Payment Status",
        "Payment Method",
        "Transaction ID / Ref",
        "Payment Verified At",
        "Verified By",
        "Events Count",
        "Events Registered",
        "Email Sent",
        "SMS Sent",
        "Admin Notes"
    ]

    for col_idx, header in enumerate(headers_1, 1):
        cell = ws1.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border = thin_border
    ws1.row_dimensions[4].height = 28

    current_row = 5
    for reg in qs:
        p = getattr(reg, 'participant', None)
        events_str = ", ".join([
            f"{re.event.name} ({'Team: ' + re.team_name if re.is_team and re.team_name else 'Individual'})"
            for re in reg.registration_events.all() if re.event
        ])
        verified_by_str = reg.payment_verified_by.username if reg.payment_verified_by else "-"

        student_name = p.name if p else "N/A"
        student_email = p.email if p else "N/A"
        student_phone = p.phone if p else "N/A"
        student_school = p.school_display if p else "N/A"
        student_grade = (p.get_grade_display() if hasattr(p, 'get_grade_display') else p.grade) if p else "-"
        student_group = (f"Group {p.group.code}" if (p and p.group) else (f"Grade {p.grade}" if p else "-"))

        row_data = [
            reg.short_code,
            str(reg.confirmation_code),
            timezone.localtime(reg.registered_at).strftime('%Y-%m-%d %H:%M') if reg.registered_at else "-",
            student_name,
            student_email,
            student_phone,
            student_school,
            student_grade,
            student_group,
            reg.total_fee,
            reg.payment_status,
            reg.get_payment_method_display() if hasattr(reg, 'get_payment_method_display') else reg.payment_method,
            reg.payment_reference or "-",
            timezone.localtime(reg.payment_verified_at).strftime('%Y-%m-%d %H:%M') if reg.payment_verified_at else "-",
            verified_by_str,
            reg.registration_events.count(),
            events_str,
            "Yes" if reg.email_sent else "No",
            "Yes" if reg.sms_sent else "No",
            reg.admin_notes or "-"
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=current_row, column=col_idx, value=value)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            # Formatting specific columns
            if col_idx in (1, 2, 3, 8, 9, 14, 15, 16, 18, 19):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 10:  # Fee
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '৳ #,##0'
                cell.font = bold_font
            elif col_idx == 11:  # Status
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if reg.payment_status in status_styles:
                    cell.fill = status_styles[reg.payment_status]['fill']
                    cell.font = status_styles[reg.payment_status]['font']

        ws1.row_dimensions[current_row].height = 20
        current_row += 1

    # -------------------------------------------------------------
    # TAB 2: EVENTS & TEAM PARTICIPANT ROSTER
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Competitions Roster")
    ws2.views.sheetView[0].showGridLines = True

    ws2['A1'] = "JOSEPHITE TECH CLUB — 17 COMPETITIONS PARTICIPANT ROSTER"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Generated: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')}"
    ws2['A2'].font = meta_font

    headers_2 = [
        "Event ID",
        "Competition Name",
        "Category",
        "Reg ID",
        "Student Name",
        "Email",
        "Phone",
        "School / Institution",
        "Grade / Group",
        "Entry Type",
        "Team Name",
        "Team Members Roster",
        "Fee Charged (BDT)",
        "Payment Status",
        "Payment Reference"
    ]

    for col_idx, header in enumerate(headers_2, 1):
        cell = ws2.cell(row=4, column=col_idx, value=header)
        cell.fill = sub_header_fill
        cell.font = sub_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws2.row_dimensions[4].height = 26

    current_row_2 = 5
    reg_events = RegistrationEvent.objects.select_related(
        'event', 'registration', 'registration__participant', 'registration__participant__school'
    )
    if qs is not None:
        reg_events = reg_events.filter(registration__in=qs)
    reg_events = reg_events.order_by('event__order', 'event__name', '-registration__registered_at')

    for re in reg_events:
        r = getattr(re, 'registration', None)
        if not r:
            continue
        p = getattr(r, 'participant', None)
        ev = getattr(re, 'event', None)
        if not ev:
            continue

        student_name = p.name if p else "N/A"
        student_email = p.email if p else "N/A"
        student_phone = p.phone if p else "N/A"
        student_school = p.school_display if p else "N/A"
        student_grade = (p.get_grade_display() if hasattr(p, 'get_grade_display') else p.grade) if p else "-"

        row_data_2 = [
            ev.id,
            ev.name,
            ev.get_category_display() if hasattr(ev, 'get_category_display') else ev.category,
            r.short_code,
            student_name,
            student_email,
            student_phone,
            student_school,
            student_grade,
            "Team" if re.is_team else "Individual",
            re.team_name or "-",
            re.team_members or "-",
            re.fee_charged,
            r.payment_status,
            r.payment_reference or "-"
        ]

        for col_idx, value in enumerate(row_data_2, 1):
            cell = ws2.cell(row=current_row_2, column=col_idx, value=value)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            if col_idx in (1, 4, 9, 10):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 13:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = '৳ #,##0'
            elif col_idx == 14:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if r.payment_status in status_styles:
                    cell.fill = status_styles[r.payment_status]['fill']
                    cell.font = status_styles[r.payment_status]['font']

        ws2.row_dimensions[current_row_2].height = 19
        current_row_2 += 1

    # -------------------------------------------------------------
    # TAB 3: EXECUTIVE & FINANCIAL SUMMARY
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Financial & Event Summary")
    ws3.views.sheetView[0].showGridLines = True

    ws3['A1'] = "EXECUTIVE SUMMARY & EVENT REGISTRATION METRICS"
    ws3['A1'].font = title_font
    ws3['A2'] = f"Generated: {timezone.localtime().strftime('%B %d, %Y at %I:%M %p')}"
    ws3['A2'].font = meta_font

    # Overview Table
    headers_3 = ["Metric / Status", "Count", "Total Value (BDT)"]
    for col_idx, header in enumerate(headers_3, 1):
        cell = ws3.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    from django.db.models import Count, Sum
    verified_cnt = Registration.objects.filter(payment_status='VERIFIED').count()
    verified_rev = Registration.objects.filter(payment_status='VERIFIED').aggregate(s=Sum('total_fee'))['s'] or 0

    pending_cnt = Registration.objects.filter(payment_status='PENDING').count()
    pending_rev = Registration.objects.filter(payment_status='PENDING').aggregate(s=Sum('total_fee'))['s'] or 0

    rejected_cnt = Registration.objects.filter(payment_status='REJECTED').count()
    rejected_rev = Registration.objects.filter(payment_status='REJECTED').aggregate(s=Sum('total_fee'))['s'] or 0

    total_cnt = Registration.objects.count()
    total_rev = Registration.objects.aggregate(s=Sum('total_fee'))['s'] or 0

    summary_rows = [
        ("Verified Registrations (Paid)", verified_cnt, verified_rev, 'VERIFIED'),
        ("Pending Verification", pending_cnt, pending_rev, 'PENDING'),
        ("Rejected / Disqualified", rejected_cnt, rejected_rev, 'REJECTED'),
        ("TOTAL REGISTERED SUBMISSIONS", total_cnt, total_rev, None),
    ]

    for idx, (label, count, val, st_key) in enumerate(summary_rows, 5):
        c1 = ws3.cell(row=idx, column=1, value=label)
        c2 = ws3.cell(row=idx, column=2, value=count)
        c3 = ws3.cell(row=idx, column=3, value=val)

        c1.font = bold_font if idx == 8 else regular_font
        c2.font = bold_font
        c3.font = bold_font

        c2.alignment = Alignment(horizontal='center', vertical='center')
        c3.alignment = Alignment(horizontal='right', vertical='center')
        c3.number_format = '৳ #,##0'

        for c in (c1, c2, c3):
            c.border = thin_border
            if st_key and st_key in status_styles:
                c.fill = status_styles[st_key]['fill']

    # Per-Event Stats Table
    ws3['A11'] = "COMPETITIONS ENROLLMENT BREAKDOWN"
    ws3['A11'].font = Font(name='Arial', size=12, bold=True, color='010B1E')

    ev_headers = ["Event Name", "Category", "Total Contestants", "Team Submissions", "Total Fees Collected (BDT)"]
    for col_idx, header in enumerate(ev_headers, 1):
        cell = ws3.cell(row=13, column=col_idx, value=header)
        cell.fill = sub_header_fill
        cell.font = sub_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    events = Event.objects.all().order_by('order', 'name')
    for row_idx, ev in enumerate(events, 14):
        ev_regs = RegistrationEvent.objects.filter(event=ev)
        cnt = ev_regs.count()
        team_cnt = ev_regs.filter(is_team=True).count()
        fees = ev_regs.filter(registration__payment_status='VERIFIED').aggregate(s=Sum('fee_charged'))['s'] or 0

        ws3.cell(row=row_idx, column=1, value=ev.name).border = thin_border
        ws3.cell(row=row_idx, column=2, value=ev.get_category_display() if hasattr(ev, 'get_category_display') else ev.category).border = thin_border
        
        c3 = ws3.cell(row=row_idx, column=3, value=cnt)
        c3.alignment = Alignment(horizontal='center')
        c3.border = thin_border
        
        c4 = ws3.cell(row=row_idx, column=4, value=team_cnt)
        c4.alignment = Alignment(horizontal='center')
        c4.border = thin_border
        
        c5 = ws3.cell(row=row_idx, column=5, value=fees)
        c5.alignment = Alignment(horizontal='right')
        c5.number_format = '৳ #,##0'
        c5.border = thin_border

    # Auto-adjust column widths on all worksheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Ignore merged cells or title in calculation
                if cell.row < 4 and col_letter == 'A':
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 48)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
